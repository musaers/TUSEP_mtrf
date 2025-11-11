from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime, timezone
from io import BytesIO
from typing import List, Dict, Any
from sqlalchemy.orm import Session
from database import Device, FaultRecord

class ExcelReportService:
    """TÜSEP Excel Rapor Oluşturma Servisi - PostgreSQL Version"""
    
    @staticmethod
    def create_workbook():
        wb = Workbook()
        ws = wb.active
        return wb, ws
    
    @staticmethod
    def style_header(ws, row, start_col, end_col, title, bg_color="366092"):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws.cell(row=row, column=start_col).value = title
        if end_col > start_col:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    
    @staticmethod
    def format_time_minutes(hours: float) -> str:
        if hours == 0:
            return "00:00"
        total_seconds = int(hours * 3600)
        minutes = total_seconds // 60
        seconds = total_seconds % 60
        return f"{minutes:02d}:{seconds:02d}"
    
    @staticmethod
    async def generate_device_failure_frequency_report_postgres(db: Session, year: int = None):
        """Gİ.YD.DH.08 - PostgreSQL version"""
        if not year:
            year = datetime.now(timezone.utc).year
        
        wb, ws = ExcelReportService.create_workbook()
        ws.title = f"Arızalanma Sıklığı {year}"
        
        ExcelReportService.style_header(ws, 1, 1, 15, f"Gİ.YD.DH.08 - CİHAZ ARIZALANMA SIKLIĞI - {year}")
        
        headers = ["Cihaz Kodu", "Cihaz Tipi", "Konum", "OCAK", "ŞUBAT", "MART", 
                   "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ", "AĞUSTOS", "EYLÜL",
                   "EKİM", "KASIM", "ARALIK", "YILLIK TOPLAM"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        devices = db.query(Device).all()
        
        row_num = 3
        for device in devices:
            ws.cell(row=row_num, column=1).value = device.code
            ws.cell(row=row_num, column=2).value = device.type
            ws.cell(row=row_num, column=3).value = device.location
            
            monthly_counts = [0] * 12
            faults = db.query(FaultRecord).filter(FaultRecord.device_id == device.id).all()
            
            for fault in faults:
                if fault.created_at and fault.created_at.year == year:
                    month_index = fault.created_at.month - 1
                    monthly_counts[month_index] += 1
            
            for i, count in enumerate(monthly_counts):
                ws.cell(row=row_num, column=4 + i).value = count
            
            ws.cell(row=row_num, column=16).value = sum(monthly_counts)
            ws.cell(row=row_num, column=16).font = Font(bold=True)
            
            row_num += 1
        
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
        for col in range(4, 17):
            ws.column_dimensions[chr(64 + col)].width = 12
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    async def generate_intervention_duration_report_postgres(db: Session, year: int = None):
        """Gİ.YD.DH.07 - PostgreSQL version"""
        if not year:
            year = datetime.now(timezone.utc).year
        
        wb, ws = ExcelReportService.create_workbook()
        ws.title = f"Müdahale Süresi {year}"
        
        ExcelReportService.style_header(ws, 1, 1, 7, f"Gİ.YD.DH.07 - CİHAZ ARIZALARINA MÜDAHALE SÜRESİ - {year}")
        
        headers = ["Bölüm/Lokasyon", "1. ÇEYREK\n(Ocak-Mart)", "2. ÇEYREK\n(Nisan-Haziran)", 
                   "3. ÇEYREK\n(Temmuz-Eylül)", "4. ÇEYREK\n(Ekim-Aralık)", 
                   "YILLIK ORTALAMA", "TOPLAM ARIZA"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        ws.row_dimensions[2].height = 40
        
        locations = {}
        devices = db.query(Device).all()
        
        for device in devices:
            location = device.location
            if location not in locations:
                locations[location] = {
                    'q1_durations': [],
                    'q2_durations': [],
                    'q3_durations': [],
                    'q4_durations': [],
                    'total_faults': 0
                }
            
            faults = db.query(FaultRecord).filter(
                FaultRecord.device_id == device.id,
                FaultRecord.status == "closed"
            ).all()
            
            for fault in faults:
                if fault.created_at and fault.created_at.year == year:
                    month = fault.created_at.month
                    duration = fault.repair_duration
                    
                    locations[location]['total_faults'] += 1
                    
                    if month in [1, 2, 3]:
                        locations[location]['q1_durations'].append(duration)
                    elif month in [4, 5, 6]:
                        locations[location]['q2_durations'].append(duration)
                    elif month in [7, 8, 9]:
                        locations[location]['q3_durations'].append(duration)
                    elif month in [10, 11, 12]:
                        locations[location]['q4_durations'].append(duration)
        
        row_num = 3
        for location, data in locations.items():
            ws.cell(row=row_num, column=1).value = location
            
            q1_avg = sum(data['q1_durations']) / len(data['q1_durations']) if data['q1_durations'] else 0
            q2_avg = sum(data['q2_durations']) / len(data['q2_durations']) if data['q2_durations'] else 0
            q3_avg = sum(data['q3_durations']) / len(data['q3_durations']) if data['q3_durations'] else 0
            q4_avg = sum(data['q4_durations']) / len(data['q4_durations']) if data['q4_durations'] else 0
            
            ws.cell(row=row_num, column=2).value = ExcelReportService.format_time_minutes(q1_avg)
            ws.cell(row=row_num, column=3).value = ExcelReportService.format_time_minutes(q2_avg)
            ws.cell(row=row_num, column=4).value = ExcelReportService.format_time_minutes(q3_avg)
            ws.cell(row=row_num, column=5).value = ExcelReportService.format_time_minutes(q4_avg)
            
            all_durations = data['q1_durations'] + data['q2_durations'] + data['q3_durations'] + data['q4_durations']
            yearly_avg = sum(all_durations) / len(all_durations) if all_durations else 0
            ws.cell(row=row_num, column=6).value = ExcelReportService.format_time_minutes(yearly_avg)
            ws.cell(row=row_num, column=6).font = Font(bold=True)
            
            ws.cell(row=row_num, column=7).value = data['total_faults']
            
            row_num += 1
        
        ws.column_dimensions['A'].width = 30
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @staticmethod
    async def generate_facility_issues_report_postgres(db: Session, year: int = None):
        """Gİ.YD.DH.02 - PostgreSQL version"""
        if not year:
            year = datetime.now(timezone.utc).year
        
        wb, ws = ExcelReportService.create_workbook()
        ws.title = f"Tesis Sorunları {year}"
        
        ExcelReportService.style_header(ws, 1, 1, 14, f"Gİ.YD.DH.02 - TESİS KAYNAKLI SORUNLARA MÜDAHALE SÜRESİ - {year}")
        
        headers = ["Ay", "Tesis Kaynaklı Toplam Arıza", "Müdahale Süresi (dk:sn)", "Ortalama Müdahale"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        months_tr = ["OCAK", "ŞUBAT", "MART", "NİSAN", "MAYIS", "HAZİRAN",
                     "TEMMUZ", "AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK"]
        
        row_num = 3
        yearly_total_faults = 0
        yearly_total_duration = 0
        
        for month_num, month_name in enumerate(months_tr, 1):
            faults = db.query(FaultRecord).filter(FaultRecord.status == "closed").all()
            
            month_faults = []
            for fault in faults:
                if fault.created_at and fault.created_at.year == year and fault.created_at.month == month_num:
                    desc = fault.description.lower()
                    if 'tesis' in desc or 'altyapı' in desc or 'elektrik' in desc:
                        month_faults.append(fault)
            
            fault_count = len(month_faults)
            total_duration = sum(f.repair_duration for f in month_faults)
            avg_duration = total_duration / fault_count if fault_count > 0 else 0
            
            ws.cell(row=row_num, column=1).value = month_name
            ws.cell(row=row_num, column=2).value = fault_count
            ws.cell(row=row_num, column=3).value = ExcelReportService.format_time_minutes(total_duration)
            ws.cell(row=row_num, column=4).value = ExcelReportService.format_time_minutes(avg_duration)
            
            yearly_total_faults += fault_count
            yearly_total_duration += total_duration
            
            row_num += 1
        
        ws.cell(row=row_num, column=1).value = "YILLIK TOPLAM"
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=2).value = yearly_total_faults
        ws.cell(row=row_num, column=2).font = Font(bold=True)
        ws.cell(row=row_num, column=3).value = ExcelReportService.format_time_minutes(yearly_total_duration)
        ws.cell(row=row_num, column=3).font = Font(bold=True)
        
        yearly_avg = yearly_total_duration / yearly_total_faults if yearly_total_faults > 0 else 0
        ws.cell(row=row_num, column=4).value = ExcelReportService.format_time_minutes(yearly_avg)
        ws.cell(row=row_num, column=4).font = Font(bold=True)
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
